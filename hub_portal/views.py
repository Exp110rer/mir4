from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, TemplateView
from django.http import HttpResponse, HttpResponseRedirect, HttpRequest
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from hub_api.models import Order, Composition, Item, ProductCategory
from mirusers.models import Hub
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font
import io
from django.db.models import F


# Create your views here.

def get_filter_query(request, option=None):
    filter_query = dict()
    if option == 'fdso':
        filter_query['status'] = 2
        filter_query['deleted'] = False
        filter_query['csDownloadStatus'] = True
        return filter_query
    f_date = request.GET.get('filter_date', None)
    filter_date = datetime.strptime(f_date, "%Y-%m-%d") if f_date else None
    filter_productCategory = request.GET.get('filter_productCategory', 'ALL')
    filter_hub = request.GET.get('filter_hub', 'ALL')
    if filter_date:
        filter_query['buyoutDate'] = filter_date
    else:
        filter_query['buyoutDate__gte'] = datetime.now()
        filter_query['buyoutDate__lt'] = datetime.now() + timedelta(days=14)
    if filter_productCategory != 'ALL':
        filter_query['productCategory'] = filter_productCategory
    if filter_hub != 'ALL':
        filter_query['hub__name'] = filter_hub
    if option == 'csValidation':
        filter_query['status'] = 2
        filter_query['deleted'] = False
        filter_query['csDownloadStatus'] = False
    if option == 'fdsoNumberQuery':
        filter_query['status'] = 2
        filter_query['deleted'] = False
        filter_query['csDownloadStatus'] = True
    return filter_query


class OrderBCPListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Order
    template_name = 'hub_portal/orders_bcp.html'

    # queryset = Order.objects.filter(buyoutDate__gte=datetime.now(), buyoutDate__lt=datetime.now() + timedelta(days=7))

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists() or self.request.user.groups.filter(
            name='Orders_BCP').exists()

    def get_context_data(self, *, object_list=None, **kwargs):
        filter_date = self.request.GET.get('filter_date', None)
        filter_productCategory = self.request.GET.get('filter_productCategory', 'ALL')
        filter_hub = self.request.GET.get('filter_hub', 'ALL')
        context = super().get_context_data()
        if self.request.user.groups.filter(name='Orders_CS').exists():
            context['user_group'] = 'CS'
        elif self.request.user.groups.filter(name='Orders_BCP').exists():
            context['user_group'] = 'BCP'
        else:
            context['user_group'] = 'NA'
        context['product_categories'] = ProductCategory.objects.all()
        context['hubs'] = Hub.objects.filter(id__gt=1)
        context['filter_date'] = filter_date
        context['filter_productCategory'] = filter_productCategory
        context['filter_hub'] = filter_hub
        return context

    def dispatch(self, request, *args, **kwargs):
        f_date = request.GET.get('filter_date', None)
        filter_date = datetime.strptime(f_date, "%Y-%m-%d") if f_date else None
        filter_productCategory = request.GET.get('filter_productCategory', 'ALL')
        filter_hub = request.GET.get('filter_hub', 'ALL')
        filter_query = dict()
        if filter_date:
            filter_query['buyoutDate'] = filter_date
        else:
            filter_query['buyoutDate__gte'] = datetime.now()
            filter_query['buyoutDate__lt'] = datetime.now() + timedelta(days=14)
        if filter_productCategory != 'ALL':
            filter_query['productCategory'] = filter_productCategory
        if filter_hub != 'ALL':
            filter_query['hub__name'] = filter_hub
        print("Filter query", filter_query)
        self.queryset = Order.objects.filter(**filter_query).order_by('buyoutDate')
        return super().dispatch(request)


class OrderCSListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Order
    template_name = 'hub_portal/orders.html'

    # queryset = Order.objects.filter(buyoutDate__gte=datetime.now(), buyoutDate__lt=datetime.now() + timedelta(days=7))

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists() or self.request.user.groups.filter(
            name='Orders_BCP').exists()

    def get_context_data(self, *, object_list=None, **kwargs):
        filter_date = self.request.GET.get('filter_date', None)
        filter_productCategory = self.request.GET.get('filter_productCategory', 'ALL')
        filter_hub = self.request.GET.get('filter_hub', 'ALL')
        context = super().get_context_data()
        if self.request.user.groups.filter(name='Orders_CS').exists():
            context['user_group'] = 'CS'
        elif self.request.user.groups.filter(name='Orders_BCP').exists():
            context['user_group'] = 'BCP'
        else:
            context['user_group'] = 'NA'
        context['product_categories'] = ProductCategory.objects.all()
        context['hubs'] = Hub.objects.filter(id__gt=1)
        context['filter_date'] = filter_date
        context['filter_productCategory'] = filter_productCategory
        context['filter_hub'] = filter_hub
        context['csOrdersForValidation'] = Order.objects.filter(
            **get_filter_query(self.request, option='csValidation')).count()
        context['csOrdersFor1CReadiness'] = Order.objects.filter(status=2, deleted=False, csDownloadStatus=1,
                                                                 csValidityStatus=1, csDownloadedBy=self.request.user,
                                                                 cs1CReadinessStatus=0).count()
        return context

    def dispatch(self, request, *args, **kwargs):
        filter_query = get_filter_query(request)
        self.queryset = Order.objects.filter(**filter_query).order_by('buyoutDate')
        return super().dispatch(request)


class OrderClientBasedListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists()

    def dispatch(self, request, *args, **kwargs):
        print('Startik')
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()
        filter_query = get_filter_query(request, option='csValidation')
        self.queryset = Order.objects.filter(**filter_query)
        orders = list(Order.objects.filter(**filter_query))
        if not orders or not filter_query['buyoutDate'] or filter_query['hub__name'] == 'ALL':
            return HttpResponseRedirect(self.request.META.get('HTTP_REFERER'))
        filter_query_fdsoNumberQuery = get_filter_query(request, option='fdsoNumberQuery')
        _order_fdso_number_update = Order.objects.only('csFDSOfileNumber').filter(**filter_query_fdsoNumberQuery)
        if _order_fdso_number_update.first():
            csFDSOfileNumber = _order_fdso_number_update.first().csFDSOfileNumber
        else:
            csFDSOfileNumber = 0
        self.queryset.update(csDownloadedBy=self.request.user, csValidityStatus=True, csDownloadStatus=True)
        _order_fdso_number_update.update(csFDSOfileNumber=csFDSOfileNumber + 1)
        return HttpResponseRedirect(self.request.META.get('HTTP_REFERER'))


class OrderBCPDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    def dispatch(self, request, pk, *args, **kwargs):
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()
        order = Order.objects.get(id=pk)
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('TDSheet', 0)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 10
        items = Item.objects.filter(composition__order=order, validity=1)
        ws.cell(row=1, column=2, value='gtin')
        ws.cell(row=1, column=3, value='sku')
        ws.cell(row=1, column=4, value='batch')
        for data in enumerate(items, start=2):
            ws.cell(row=data[0], column=2, value=data[1].tuid)
            ws.cell(row=data[0], column=3, value=f"0000000000{data[1].sku}")
            batch = f"{data[1].batch}" if data[1].batch else ''
            ws.cell(row=data[0], column=4, value=batch)
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_name = f'filename="{order.order}.xlsx"'
        return HttpResponse(file_obj.getvalue(), headers={'Content-Type': 'application/vnd.ms-excel',
                                                          'Content-Disposition': f"inline; {file_name}"})

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_BCP').exists()


class OrderCSDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    def dispatch(self, request, pk, *args, **kwargs):
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()
        order = Order.objects.get(id=pk)
        if order.status != 2:
            return self.handle_no_permission()
        order.downloadedBy_id = request.user.id
        order.save(update_fields=['downloadedBy_id'])
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('')
        if "Sheet" in wb.sheetnames:
            wb.remove(wb['Sheet'])
        values = ['PurchaseOrderNumber',
                  'Delivery Date',
                  'Material',
                  'Quantity',
                  'Unit of Measure',
                  'Batch',
                  'Owner']
        for i, value in enumerate(values, start=1):
            ws.cell(row=1, column=i, value=value)
            ws.column_dimensions[chr(i + 64)].width = 20
        compositions = Composition.objects.filter(order=order)
        i = 2
        for composition in compositions:
            composition_items = Item.objects.filter(composition=composition)
            composition_batches = set([item.batch for item in composition_items])
            if not composition_batches: composition_batches = {None}
            for composition_batch in composition_batches:
                ws.cell(row=i, column=1, value=order.order)
                ws.cell(row=i, column=2, value=order.buyoutDate)
                ws.cell(row=i, column=3, value=composition.sku)
                if order.traceability:
                    initial_amount = len(composition_items.filter(batch=composition_batch))
                    validated_amount = len(composition_items.filter(batch=composition_batch, validity=1))
                    ws.cell(row=i, column=4, value=validated_amount)
                    if initial_amount != validated_amount:
                        ws.cell(row=i, column=4).font = Font(color='FF0000')
                else:
                    ws.cell(row=i, column=4, value=composition.amount)
                if composition.unitOfMeasure == 'case':
                    ws.cell(row=i, column=5, value='CS')
                elif composition.unitOfMeasure == 'out':
                    ws.cell(row=i, column=5, value='OUT')
                else:
                    ws.cell(row=i, column=5, value='Unknown')
                ws.cell(row=i, column=6, value=composition_batch)
                ws.cell(row=i, column=7, value=order.saleType)
                i += 1
        file_obj = io.BytesIO()
        wb.save(file_obj)
        buyoutDate = str(order.buyoutDate).replace('-', '')
        if order.contractType == 't':
            file_contact_type = 's'
        elif order.contractType == 'c':
            file_contact_type = 'k'
        file_name = f'filename="{order.order}_{order.hub.name}_Sales_{order.productCategory}_{buyoutDate}_{order.saleType}_{file_contact_type}.xlsx"'
        # order.downloadedBy_id = request.user.id
        # order.save()
        return HttpResponse(file_obj.getvalue(), headers={'Content-Type': 'application/vnd.ms-excel',
                                                          'Content-Disposition': f"inline; {file_name}"})

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists()


class OrderFDSODetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    def dispatch(self, request, pk, *args, **kwargs):
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()
        filter_query = get_filter_query(request, option='fdso')
        _order = Order.objects.only('buyoutDate', 'hub', 'csFDSOfileNumber').get(pk=pk)
        _buyoutDate = _order.buyoutDate
        _hub = _order.hub
        _fdsoFileNumber = _order.csFDSOfileNumber
        orders = list(Order.objects.filter(buyoutDate=_buyoutDate, hub=_hub, **filter_query))
        if not orders:
            return HttpResponseRedirect(self.request.META.get('HTTP_REFERER'))
        wb = openpyxl.Workbook()
        ws = wb.create_sheet('')
        if "Sheet" in wb.sheetnames:
            wb.remove(wb['Sheet'])
        values = ['CustomerOrderNumber',
                  'Hub',
                  'Stock type',
                  'PurchaseOrderNumber',
                  'Delivery Date',
                  'Material',
                  'Quantity',
                  'Unit of Measure',
                  'Batch',
                  'Owner']
        for i, value in enumerate(values, start=1):
            ws.cell(row=1, column=i, value=value)
            ws.column_dimensions[chr(i + 64)].width = 20
        i = 2
        for order in orders:
            if order.traceability:
                items = list(Item.objects.only('sku', 'batch', 'unitOfMeasure').filter(validity=True, composition__order_id=order.id))
                items_all = list(Item.objects.only('sku', 'batch', 'unitOfMeasure').filter(composition__order_id=order.id))
                skus = sorted(list(set(item.sku for item in items)))
                for sku in skus:
                    batches = set([item.batch for item in items if item.sku == sku])
                    for batch in batches:
                        ws.cell(row=i, column=1, value=order.customerOrderNumber)
                        ws.cell(row=i, column=2, value=order.hub.name)
                        if order.contractType == 't':
                            ws.cell(row=i, column=3, value='s')
                        elif order.contractType == 'c':
                            ws.cell(row=i, column=3, value='k')
                        ws.cell(row=i, column=4, value=order.order)
                        ws.cell(row=i, column=5, value=order.buyoutDate)
                        ws.cell(row=i, column=6, value=sku)
                        items_len = len([item for item in items if item.sku == sku and item.batch == batch])
                        items_all_len = len([item for item in items_all if item.sku == sku and item.batch == batch])
                        ws.cell(row=i, column=7,
                                value=items_len)
                        if items_len != items_all_len:
                            ws.cell(row=i, column=7).font = Font(color='FF0000')
                        if items[0].unitOfMeasure == 'case':
                            ws.cell(row=i, column=8, value='CS')
                        elif items[0].unitOfMeasure == 'out':
                            ws.cell(row=i, column=8, value='OUT')
                        else:
                            ws.cell(row=i, column=8, value='Unknown')
                        ws.cell(row=i, column=9, value=batch)
                        ws.cell(row=i, column=10, value=order.saleType)
                        i += 1
            else:
                compositions = Composition.objects.only('sku', 'amount').filter(order=order)
                for composition in compositions:
                    ws.cell(row=i, column=1, value=order.customerOrderNumber)
                    ws.cell(row=i, column=2, value=order.hub.name)
                    if order.contractType == 't':
                        ws.cell(row=i, column=3, value='s')
                    elif order.contractType == 'c':
                        ws.cell(row=i, column=3, value='k')
                    ws.cell(row=i, column=4, value=order.order)
                    ws.cell(row=i, column=5, value=order.buyoutDate)
                    ws.cell(row=i, column=6, value=composition.sku)
                    ws.cell(row=i, column=7, value=composition.amount)
                    if composition.unitOfMeasure == 'case':
                        ws.cell(row=i, column=8, value='CS')
                    elif composition.unitOfMeasure == 'out':
                        ws.cell(row=i, column=8, value='OUT')
                    else:
                        ws.cell(row=i, column=8, value='Unknown')
                    ws.cell(row=i, column=10, value=order.saleType)
                    i += 1

        # for order in orders:
        #     if order.status != 2:
        #         return self.handle_no_permission()
        #     compositions = Composition.objects.filter(order=order)
        #     for composition in compositions:
        #         composition_items = Item.objects.filter(composition=composition)
        #         composition_batches = set([item.batch for item in composition_items])
        #         if not composition_batches: composition_batches = {None}
        #         for composition_batch in composition_batches:
        #             ws.cell(row=i, column=1, value=order.customerOrderNumber)
        #             ws.cell(row=i, column=2, value=order.hub.name)
        #             if order.contractType == 't':
        #                 ws.cell(row=i, column=3, value='s')
        #             elif order.contractType == 'c':
        #                 ws.cell(row=i, column=3, value='k')
        #             ws.cell(row=i, column=4, value=order.order)
        #             ws.cell(row=i, column=5, value=order.buyoutDate)
        #             ws.cell(row=i, column=6, value=composition.sku)
        #             if order.traceability:
        #                 initial_amount = len(composition_items.filter(batch=composition_batch))
        #                 validated_amount = len(composition_items.filter(batch=composition_batch, validity=1))
        #                 ws.cell(row=i, column=7, value=validated_amount)
        #                 if initial_amount != validated_amount:
        #                     ws.cell(row=i, column=7).font = Font(color='FF0000')
        #             else:
        #                 ws.cell(row=i, column=7, value=composition.amount)
        #             if composition.unitOfMeasure == 'case':
        #                 ws.cell(row=i, column=8, value='CS')
        #             elif composition.unitOfMeasure == 'out':
        #                 ws.cell(row=i, column=8, value='OUT')
        #             else:
        #                 ws.cell(row=i, column=8, value='Unknown')
        #             ws.cell(row=i, column=9, value=composition_batch)
        #             ws.cell(row=i, column=10, value=order.saleType)
        #             i += 1
        file_obj = io.BytesIO()
        wb.save(file_obj)
        file_name = f'filename="FDSO_{_buyoutDate}_{_hub.name}_{_fdsoFileNumber}.xlsx"'
        return HttpResponse(file_obj.getvalue(), headers={'Content-Type': 'application/vnd.ms-excel',
                                                          'Content-Disposition': f"inline; {file_name}"})

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists()


def csValidityStatus_change(request, pk):
    try:
        order = Order.objects.get(pk=pk)
    except Order.DoesNotExist:
        return HttpResponse('Order does not exist')
    else:
        if request.user == order.csDownloadedBy:
            order.csValidityStatus = 0 if order.csValidityStatus == 1 else 1
            order.save()
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            return HttpResponse('User mismatch')


class ReadinessTemplateView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'hub_portal/readiness.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['orders_for_readiness'] = Order.objects.filter(status=2, deleted=False, csDownloadStatus=1,
                                                               csValidityStatus=1, csDownloadedBy=self.request.user,
                                                               cs1CReadinessStatus=0)
        filter_date = self.request.GET.get('filter_date', None)
        filter_productCategory = self.request.GET.get('filter_productCategory', 'ALL')
        filter_hub = self.request.GET.get('filter_hub', 'ALL')
        context['filter_date'] = filter_date
        context['filter_productCategory'] = filter_productCategory
        context['filter_hub'] = filter_hub
        return context

    def get(self, request, *args, **kwargs):
        user_test_result = self.get_test_func()()
        if not user_test_result:
            return self.handle_no_permission()
        _source_link = request.META.get('HTTP_REFERER', None)
        if _source_link:
            source_link = _source_link.split('/')
            if source_link[3] == 'hub_portal' and source_link[4] == 'cs':
                return super().get(request)
        return HttpResponse('ERROR: Wrong path used')

    def test_func(self):
        return self.request.user.groups.filter(name='Orders_CS').exists()


class ReadinessConfirmTemplateView(ReadinessTemplateView):
    def get(self, request, *args, **kwargs):
        context = self.get_context_data()
        try:
            context['orders_for_readiness'].update(cs1CReadinessStatus=True)
        except Exception:
            return HttpResponse('cs1CReadinessStatus update ERROR')
        finally:
            return HttpResponseRedirect(
                f"{reverse_lazy('hub_portal:orders_cs')}?filter_date={context['filter_date']}&filter_productCategory={context['filter_productCategory']}&filter_hub={context['filter_hub']}")
